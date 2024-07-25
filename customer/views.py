import csv
import json

from django.contrib import messages
from django.contrib.auth.views import LoginView
from django.contrib.sites.shortcuts import get_current_site
from django.core.mail import send_mail, EmailMessage
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.urls import reverse_lazy
from django.utils.encoding import force_str, force_bytes
from django.utils.http import urlsafe_base64_decode, urlsafe_base64_encode
from django.views import View
from django.views.generic import DetailView, FormView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.list import ListView

from customer.models import Customer, User
import openpyxl

from customer.tokens import account_activation_token


# Create your views here.
class CustomerListView(ListView):
    model = Customer
    paginate_by = 10
    template_name = 'customer/customer-list.html'
    context_object_name = 'customer_list'

    def get_queryset(self):
        searched = self.request.GET.get('searched')
        if searched:
            customer_list = Customer.objects.filter(full_name__icontains=searched)
        else:
            customer_list = Customer.objects.all()
        return customer_list

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        searched = self.request.GET.get('searched')
        context['searched'] = searched
        return context


def activate(request, uidb64, token):
    user = get_user_model()
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = user.objects.get(pk=uid)
    except:
        user = None

    if user is not None and account_activation_token.check_token(user, token):
        user.is_active = True
        user.save()

        messages.success(request, "Thank you for your email confirmation. Now you can login your account.")
        return redirect('customer:customers')
    else:
        messages.error(request, "Activation link is invalid!")

    return redirect('customer:customers')


def activateEmail(request, user, to_email):
    mail_subject = "Activate your user account."
    message = render_to_string("auth/activate_account.html", {
        'user': user.username,
        'domain': get_current_site(request).domain,
        'uid': urlsafe_base64_encode(force_bytes(user.pk)),
        'token': account_activation_token.make_token(user),
        "protocol": 'https' if request.is_secure() else 'http'
    })

    if send_mail(subject=mail_subject, message=message, from_email='shohruxabdusaidov@gmail.com',
                 recipient_list=[to_email], fail_silently=False):
        messages.success(request, f'Dear <b>{user}</b>, please go to you email <b>{to_email}</b> inbox and click on \
                received activation link to confirm and complete the registration. <b>Note:</b> Check your spam folder.')
    else:
        messages.error(request, f'Problem sending email to {to_email}, check if you typed it correctly.')


# def customers(request):
#     searched = request.GET.get('searched')
#     if searched:
#         customer_list = Customer.objects.filter(full_name__icontains=searched)
#     else:
#         customer_list = Customer.objects.all()
#         paginator = Paginator(customer_list, 4)
#         page = request.GET.get('page')
#
#         try:
#             customer_list = paginator.page(page)
#         except PageNotAnInteger:
#             customer_list = paginator.page(1)
#         except EmptyPage:
#             customer_list = paginator.page(paginator.num_pages)
#     context = {
#         'customer_list': customer_list,
#     }
#     return render(request, 'customer/customer-list.html', context)


class CustomerDetailView(DetailView):
    model = Customer
    template_name = 'customer/customer-detail.html'

    context_object_name = 'customer'


# def customer_detail(request, id):
#     customer = Customer.objects.get(id=id)
#     context = {"customer": customer}
#     return render(request, 'customer/customer-detail.html', context)


from django.contrib.auth import authenticate, login, logout, get_user_model
from django.shortcuts import render, redirect

from customer.forms import LoginForm, RegisterModelForm, CustomerModelForm, EmailForm


# def login_page(request):
# #     if request.method == 'POST':
# #         form = LoginForm(request.POST)
# #         if form.is_valid():
# #             phone_number = form.cleaned_data['phone_number']
# #             password = form.cleaned_data['password']
# #             user = authenticate(request, phone_number=phone_number, password=password)
# #             if user:
# #                 login(request, user)
# #                 return redirect('customers')
# #     else:
# #         form = LoginForm()
# #
# #     return render(request, 'auth/login.html', {'form': form})

class LoginPageView(View):

    def get(self, request):
        form = LoginForm()
        return render(request, 'auth/login.html', {'form': form})

    def post(self, request):
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            user = authenticate(request, email=email, password=password)
            if user:
                login(request, user)
                return redirect('customer:customers')

        return render(request, 'auth/login.html', {'form': form})


class LoginPage(LoginView):
    template_name = 'auth/login.html'
    redirect_authenticated_user = True
    authentication_form = LoginForm

    # success_url = reverse_lazy('customers')

    def get_success_url(self):
        return reverse_lazy('customers')


def register(request):
    if request.method == 'POST':
        form = RegisterModelForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()
            activateEmail(request, user, form.cleaned_data.get('email'))
            return redirect('customer:email_sent')

        else:
            for error in list(form.errors.values()):
                messages.error(request, error)
    else:
        form = RegisterModelForm()

    return render(request, 'auth/register.html', {"form": form})


def email_sent(request):
    return render(request, 'auth/email-sent.html', )


class RegisterFormView(FormView):
    template_name = 'auth/register.html'
    form_class = RegisterModelForm
    success_url = reverse_lazy('customer:customers')


    def form_valid(self, form):
        user = form.save(commit=False)
        user.email = form.cleaned_data['phone_number']
        user.password = form.cleaned_data['password']
        user.save()
        login(self.request, user)
        return redirect('customer:customers')


def logout_page(request):
    logout(request)
    return render(request, 'auth/logout.html')


def export_data(request):
    response = None
    format = request.GET.get('format', 'csv')
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=customers.csv'
        writer = csv.writer(response)
        writer.writerow(['Id', 'Full Name', 'Email', 'Phone Number', 'Address'])
        for customer in Customer.objects.all():
            writer.writerow([customer.id, customer.full_name, customer.email, customer.phone_number, customer.address])


    elif format == 'json':
        response = HttpResponse(content_type='application/json')
        data = list(Customer.objects.all().values('full_name', 'email', 'phone_number', 'address'))
        # response.content = json.dumps(data, indent=4)
        response.write(json.dumps(data, indent=4))
        response['Content-Disposition'] = 'attachment; filename=customers.json'
    elif format == 'xlsx':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Customers.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'My Data'

        # Write header row
        header = ['Id', 'Full Name', 'Email', 'Phone Number', 'Address']
        for col_num, column_title in enumerate(header, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Write data rows
        queryset = Customer.objects.all().values_list('id', 'full_name', 'email', 'phone_number', 'address')
        for row_num, row in enumerate(queryset, 1):
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num + 1, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response
    else:
        response = HttpResponse(status=404)
        response.content = 'Bad request'

    return response


# class AddCustomerView(View):
#     def get(self, request):
#         form = CustomerModelForm()
#         return render(request, 'customer/add-customer.html', {'form': form})
#
#     def post(self, request):
#         form = CustomerModelForm(request.POST, request.FILES)
#         if form.is_valid():
#             customer = form.save(commit=False)
#             customer.save()
#             return redirect('customers')
#         return render(request, 'customer/add-customer.html', {'form': form})

class AddCustomerView(CreateView):
    model = Customer
    form_class = CustomerModelForm
    template_name = 'customer/add-customer.html'
    success_url = reverse_lazy('customer:customers')


class UpdateCustomerView(UpdateView):
    model = Customer
    form_class = CustomerModelForm
    template_name = 'customer/update-customer.html'
    success_url = reverse_lazy('customers')


class DeleteCustomerView(View):
    def get(self, request, pk):
        customer = Customer.objects.get(pk=pk)
        customer.delete()
        return redirect('customer:customers')


def successful_email(request):
    return render(request, 'customer/successful_email.html')


def sending_email(request):
    if request.method == 'POST':
        form = EmailForm(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            message = form.cleaned_data['message']
            from_email = form.cleaned_data['from_email']
            to_email = form.cleaned_data['to_email']
            send_mail(subject, message, from_email, [to_email], fail_silently=False)
            return redirect('customer:successful_email')
    else:
        form = EmailForm()

    return render(request, 'customer/send_email.html', {'form': form})
